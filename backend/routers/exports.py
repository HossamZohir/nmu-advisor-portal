from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from core.dependencies import require_admin
from database import get_db
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

router = APIRouter(prefix="/exports", tags=["Exports"])

def style_header_row(ws, row_num, num_cols):
    red_fill = PatternFill(start_color="8B141E", end_color="8B141E", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = red_fill
        cell.font = white_font
        cell.alignment = center

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

@router.get("/registrations")
def export_registrations(user=Depends(require_admin)):
    db = get_db()

    # Get active semester
    sem = db.table("semesters").select("id, name").eq("is_active", True).execute()
    semester_name = sem.data[0]["name"] if sem.data else "No Active Semester"
    semester_id = sem.data[0]["id"] if sem.data else None

    # Get all students
    students = db.table("students").select(
        "id, student_id, name_en, name_ar, gpa, current_level, programs(name_en), users(name_en)"
    ).execute()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registration Report"

    # Title
    ws.merge_cells("A1:L1")
    ws["A1"] = f"NMU Faculty of Engineering — Registration Report — {semester_name}"
    ws["A1"].font = Font(bold=True, size=14, color="8B141E")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    # Headers
    headers = ["#", "Student ID", "Name (EN)", "Name (AR)", "Program", "Level",
               "GPA", "Advisor", "Registered Courses", "Credit Hours", "SIS", "Paid", "Status"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=header)
    style_header_row(ws, 2, len(headers))
    ws.row_dimensions[2].height = 20

    # Data
    for i, s in enumerate(students.data):
        reg = None
        courses = []
        if semester_id:
            reg_res = db.table("registrations").select(
                "id, status, sis_checked, payment_checked, total_credit_hours"
            ).eq("student_id", s["id"]).eq("semester_id", semester_id).execute()
            if reg_res.data:
                reg = reg_res.data[0]
                rc = db.table("registration_courses").select(
                    "courses(code)"
                ).eq("registration_id", reg["id"]).execute()
                courses = [r["courses"]["code"] for r in rc.data if r.get("courses")]

        program = s.get("programs", {}).get("name_en", "") if s.get("programs") else ""
        advisor = s.get("users", {}).get("name_en", "Unassigned") if s.get("users") else "Unassigned"
        status = reg["status"].capitalize() if reg else "Not Started"

        row = i + 3
        values = [
            i + 1,
            s["student_id"],
            s["name_en"],
            s["name_ar"],
            program,
            s["current_level"],
            s["gpa"],
            advisor,
            ", ".join(courses),
            reg["total_credit_hours"] if reg else 0,
            "✓" if reg and reg["sis_checked"] else "✗",
            "✓" if reg and reg["payment_checked"] else "✗",
            status,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if i % 2 == 0:
                cell.fill = PatternFill(start_color="FDF2F3", end_color="FDF2F3", fill_type="solid")

    auto_width(ws)
    ws.freeze_panes = "A3"

    # Save
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=registrations_{semester_name}.xlsx"}
    )


@router.get("/unregistered")
def export_unregistered(user=Depends(require_admin)):
    db = get_db()

    sem = db.table("semesters").select("id, name").eq("is_active", True).execute()
    semester_name = sem.data[0]["name"] if sem.data else "No Active Semester"
    semester_id = sem.data[0]["id"] if sem.data else None

    students = db.table("students").select(
        "id, student_id, name_en, name_ar, gpa, current_level, programs(name_en), users(name_en)"
    ).execute()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unregistered Students"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"NMU — Unregistered Students — {semester_name}"
    ws["A1"].font = Font(bold=True, size=14, color="8B141E")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    headers = ["#", "Student ID", "Name (EN)", "Name (AR)", "Program", "Level", "GPA", "Advisor"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=header)
    style_header_row(ws, 2, len(headers))

    row = 3
    count = 0
    for s in students.data:
        has_reg = False
        if semester_id:
            reg_res = db.table("registrations").select("id").eq(
                "student_id", s["id"]
            ).eq("semester_id", semester_id).execute()
            has_reg = len(reg_res.data) > 0

        if not has_reg:
            program = s.get("programs", {}).get("name_en", "") if s.get("programs") else ""
            advisor = s.get("users", {}).get("name_en", "Unassigned") if s.get("users") else "Unassigned"
            values = [count + 1, s["student_id"], s["name_en"], s["name_ar"],
                      program, s["current_level"], s["gpa"], advisor]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = Alignment(horizontal="center")
                if count % 2 == 0:
                    cell.fill = PatternFill(start_color="FDF2F3", end_color="FDF2F3", fill_type="solid")
            row += 1
            count += 1

    auto_width(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=unregistered_{semester_name}.xlsx"}
    )


@router.get("/advisor-summary")
def export_advisor_summary(user=Depends(require_admin)):
    db = get_db()

    sem = db.table("semesters").select("id, name").eq("is_active", True).execute()
    semester_name = sem.data[0]["name"] if sem.data else "No Active Semester"
    semester_id = sem.data[0]["id"] if sem.data else None

    advisors = db.table("users").select("id, name_en, name_ar, email").eq("role", "advisor").execute()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Advisor Summary"

    ws.merge_cells("A1:G1")
    ws["A1"] = f"NMU — Advisor Summary — {semester_name}"
    ws["A1"].font = Font(bold=True, size=14, color="8B141E")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    headers = ["#", "Advisor Name", "Email", "Total Students", "Registered", "SIS Done", "Paid"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=header)
    style_header_row(ws, 2, len(headers))

    for i, advisor in enumerate(advisors.data):
        students = db.table("students").select("id").eq("advisor_id", advisor["id"]).execute()
        student_ids = [s["id"] for s in students.data]
        total = len(student_ids)
        registered = sis = paid = 0

        if semester_id and student_ids:
            regs = db.table("registrations").select(
                "status, sis_checked, payment_checked"
            ).eq("semester_id", semester_id).in_("student_id", student_ids).execute()
            for r in regs.data:
                if r["status"] in ["submitted", "locked"]: registered += 1
                if r["sis_checked"]: sis += 1
                if r["payment_checked"]: paid += 1

        row = i + 3
        values = [i + 1, advisor["name_en"], advisor["email"], total, registered, sis, paid]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            if i % 2 == 0:
                cell.fill = PatternFill(start_color="FDF2F3", end_color="FDF2F3", fill_type="solid")

    auto_width(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=advisor_summary_{semester_name}.xlsx"}
    )